"""Simple class that allows you to easily write to multiple Excel worksheets
author: Ilias Aarab / aarabil"""


# Standard libs
import os

# External libs
import pandas as pd
import openpyxl
import plotly


class excelWriter:
    def __init__(
        self,
    ) -> None:
        """Simple excel writer to multiple worksheets. Based on openpyxl."""
        pass

    def create_excel(self, file_name: str = "file.xslx") -> None:
        """Creates a writer object and stores the accompanying empty excel file in "reports/file_name"

        Parameters
        ----------
        file_name : str, optional
            file name, by default 'file.xslx'
        """
        self.file_name = file_name
        path = "reports/"
        self.excel_path = os.path.join(path, self.file_name)
        if not os.path.isdir(path):
            os.makedirs(path)

        self.wb = openpyxl.Workbook()
        self.wb.save(self.excel_path)
        self.wb.remove(self.wb["Sheet"])  # remove default sheet
        self.pandas_writer = pd.ExcelWriter(
            self.excel_path, engine="openpyxl", mode="a"
        )

    def open_excel(self, file_name: str) -> None:
        self.wb = openpyxl.load_workbook(file_name)

    def set_ws(self, sheet_name: str):
        """sets a sheet as the main active sheet. If sheet name isn't found back, a new sheet is created.

        Parameters
        ----------
        sheet_name : str
            name of the sheet to activate.
        """
        if sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(sheet_name)

        self.active_ws = self.wb[sheet_name]
        self.active_ws_name = sheet_name

    def get_all_ws(self, names_only: bool = False):
        """Return list of worksheet names or dictionary of {name:sheet}

        Parameters
        ----------
        names_only : bool, optional
            Get names only or dict {name:sheet}, by default False

        Returns
        -------
        list/dict
            list or dict of sheetnames or {name:sheet}
        """

        if names_only:
            all_ws = self.wb.sheetnames
        else:
            all_ws = dict((ws.title, ws) for ws in self.wb.worksheets)

        return all_ws

    def add_df(self, df: pd.DataFrame, position: str = None, **kwargs):
        """adds a pandas dataframe to a worksheet

        Parameters
        ----------
        df : pd.DataFrame
            dataframe to add
        position : str, optional
            In case you append a df to an existing ws that contains data.
            Options: right/bottom, by default None

        Raises
        ------
        ValueError
            [in case position is used with startrow/startcol]
        """

        if "index" not in kwargs.values():
            index = False
        else:
            index = kwargs["index"]
        if position is not None and any(["startrow", "startcol"]) in kwargs:
            raise ValueError(
                "Cannot define position arg in conjunction with startrow/startcol args!"
            )
        if position == "right":
            kwargs["startcol"] = (
                self.active_ws.max_column + 2
            )  # leave one col empty between
        elif position == "bottom":
            kwargs["startrow"] = (
                self.active_ws.max_row + 2
            )  # leave one row empty between
        # with pd.ExcelWriter(self.excel_path, engine="openpyxl", mode="a") as writer:
        self.pandas_writer.book = self.wb
        self.pandas_writer.sheets = dict((ws.title, ws) for ws in self.wb.worksheets)
        df.to_excel(self.pandas_writer, self.active_ws_name, index=index, **kwargs)

    def add_image(
        self,
        plotly_fig: plotly.graph_objs.Figure,
        cell_anchor: str = None,
        scale=2,
    ):
        """Adds a plotly figure to an excel sheet

        Parameters
        ----------
        plotly_fig : plotly.graph_objs.Figure
            go.Figure object created with plotly
        cell_anchor : str | int, optional
            Cell anchor to position fig.
            If none fig will be placed to the right of last data column found, by default None
        scale : int, optional
            scale (quality) of image, by default 2
        """

        # Step 1: temporarily save fig
        plotly_fig.write_image("tmp.png", format="png", scale=scale)
        # Step 2: Create openpyxl image object and fix anchor
        img = openpyxl.drawing.image.Image("tmp.png")
        if not cell_anchor:
            cell_anchor = self.active_ws.max_column + 2
        if isinstance(cell_anchor, int):
            img.anchor = chr(65 + cell_anchor) + "1"
        elif isinstance(cell_anchor, str):
            img.anchor = cell_anchor
        # Step 3: Add image to worksheet
        self.active_ws.add_image(img)

    def resize_cols(self, width: str = "openpyxl bestFit") -> None:
        """Resizes columns of a worksheet.

        Parameters
        ----------
        width : str, optional
            if 'openpyxl bestFit': use bestFit method of openpyxl
            if 'element fit': changes width dependent on number of characters within longest cell of each column,
            by default "openpyxl bestFit"
        """

        column_letters = tuple(
            openpyxl.utils.get_column_letter(col_number + 1)
            for col_number in range(self.active_ws.max_column)
        )

        if width == "openpyxl bestFit":
            for column_letter in column_letters:
                self.active_ws.column_dimensions[column_letter].bestFit = True

        elif width == "element fit":
            # Step 1: find the longest value within each column
            new_widths = [0] * self.active_ws.max_column
            for eachrow in self.active_ws.iter_rows(values_only=True):
                for eachcol, eachvalue in enumerate(eachrow):
                    threshold = len(str(eachvalue))
                    if threshold > new_widths[eachcol]:
                        new_widths[eachcol] = len(str(eachvalue))
            # Step 2: update columns width with new values
            for i, column_letter in enumerate(column_letters):
                self.active_ws.column_dimensions[column_letter] = new_widths[i]

    def close(self) -> None:
        """Saves and closes worbook and writer."""
        self.pandas_writer.close()
        self.wb.save(self.excel_path)
        self.wb.close()
        os.remove("tmp.png")
